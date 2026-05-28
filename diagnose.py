#!/usr/bin/env python3
"""Diagnostico: que valores unicos hay en las celdas de asistencia"""
import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('VS.xlsx', data_only=True)
ws = wb['CONTROL']

# Recoger todos los valores unicos en las columnas de datos (5+)
valores = {}
for r in range(6, 139):
    for c in range(5, ws.max_column + 1):
        val = ws.cell(row=r, column=c).value
        if val is not None:
            val_str = str(val).strip()
            if val_str:
                if val_str not in valores:
                    valores[val_str] = 0
                valores[val_str] += 1

print("Valores unicos encontrados en celdas de asistencia:")
for v, count in sorted(valores.items(), key=lambda x: -x[1]):
    print(f"  '{v}' -> {count} veces (hex: {v.encode('utf-8').hex()})")

# Conteo de X para primeros 20 asociados
print("\n\nCONTEO DE X POR ASOCIADO (primeros 20):")
for r in range(6, 26):
    mz = ws.cell(row=r, column=2).value
    lt = ws.cell(row=r, column=3).value
    name = ws.cell(row=r, column=4).value
    x_count = 0
    other_marks = {}
    for c in range(5, ws.max_column + 1):
        val = ws.cell(row=r, column=c).value
        if val is not None:
            val_str = str(val).strip().upper()
            if val_str == 'X':
                x_count += 1
            elif val_str and val_str not in ('', ' '):
                if val_str not in other_marks:
                    other_marks[val_str] = 0
                other_marks[val_str] += 1
    lt_s = str(int(lt)) if isinstance(lt, float) and lt == int(lt) else str(lt)
    other_str = ', '.join(f"{k}={v}" for k,v in other_marks.items()) if other_marks else ''
    print(f"  {mz}-{lt_s} {name}: {x_count} X | otros: {other_str}")

# Verificar columnas 34-110 (2024-2026) vs todas
print("\n\nCONTEO X SOLO EN COLUMNAS 2024-2026 (cols 34-110):")
total_x_all = 0
total_x_target = 0
for r in range(6, 139):
    for c in range(5, ws.max_column + 1):
        val = ws.cell(row=r, column=c).value
        if val is not None and str(val).strip().upper() == 'X':
            total_x_all += 1
            if 34 <= c <= 110:
                total_x_target += 1
print(f"  Total X en todas las columnas: {total_x_all}")
print(f"  Total X en cols 34-110 (2024-2026): {total_x_target}")
print(f"  Total X en cols 5-33 (2022-2023): {total_x_all - total_x_target}")
